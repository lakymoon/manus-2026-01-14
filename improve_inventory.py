#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
库存表改进脚本
功能:
1. 修改BondDataTable,实现序号和日期自动填充
2. 根据BondDataTable自动生成销售清单(两种模板)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import defaultdict
import copy

def improve_bond_data_table(wb):
    """
    改进BondDataTable:
    1. 优化序号列公式
    2. 为出库日期列设置默认值公式
    """
    ws = wb['BondDataSheet']
    
    # 获取表格对象
    table = ws.tables['BondDataTable']
    
    print("正在改进BondDataTable...")
    
    # 遍历数据行,设置公式
    for row_idx in range(2, ws.max_row + 1):
        # 序号列(A列) - 保持原有公式
        seq_cell = ws.cell(row_idx, 1)
        if seq_cell.value is None or (isinstance(seq_cell.value, str) and seq_cell.value.startswith('=')):
            seq_cell.value = f'=ROW(BondDataTable[[#This Row],[序号]])-1'
        
        # 出库日期列(B列) - 如果为空,自动填充今天日期
        date_cell = ws.cell(row_idx, 2)
        if date_cell.value is None:
            date_cell.value = date.today()
            date_cell.number_format = 'YYYY-MM-DD'
        elif isinstance(date_cell.value, datetime):
            date_cell.number_format = 'YYYY-MM-DD'
        
        # 净重列(G列) - 确保公式正确
        net_weight_cell = ws.cell(row_idx, 7)
        if net_weight_cell.value is None or (isinstance(net_weight_cell.value, str) and net_weight_cell.value.startswith('=')):
            net_weight_cell.value = f'=BondDataTable[[#This Row],[毛重]]-BondDataTable[[#This Row],[除皮]]'
    
    print("✓ BondDataTable改进完成")
    return ws

def read_bond_data(ws):
    """
    从BondDataSheet读取数据
    返回: 数据列表,每行为一个字典
    """
    data = []
    
    # 从第2行开始读取(第1行是表头)
    for row_idx in range(2, ws.max_row + 1):
        # 读取净重单元格,如果是公式则计算值
        net_weight_cell = ws.cell(row_idx, 7)
        if net_weight_cell.data_type == 'f':  # 如果是公式
            # 尝试获取计算后的值
            try:
                net_weight = float(ws.cell(row_idx, 5).value) - float(ws.cell(row_idx, 6).value)
            except:
                net_weight = 0.0
        else:
            net_weight = net_weight_cell.value
        
        row_data = {
            '序号': ws.cell(row_idx, 1).value,
            '出库日期': ws.cell(row_idx, 2).value,
            '规格': ws.cell(row_idx, 3).value,
            '个数': ws.cell(row_idx, 4).value,
            '毛重': ws.cell(row_idx, 5).value,
            '除皮': ws.cell(row_idx, 6).value,
            '净重': net_weight,
            '出库对象': ws.cell(row_idx, 8).value,
            '入账': ws.cell(row_idx, 9).value,
            '备注': ws.cell(row_idx, 10).value,
            'row_idx': row_idx  # 记录行号,用于后续标记
        }
        
        # 跳过空行
        if row_data['出库日期'] is None or row_data['出库对象'] is None:
            continue
        
        # 处理日期格式
        if isinstance(row_data['出库日期'], datetime):
            row_data['出库日期'] = row_data['出库日期'].date()
        
        data.append(row_data)
    
    return data

def group_data_by_date_and_customer(data):
    """
    按出库日期和出库对象分组
    返回: {(日期, 客户): [数据列表]}
    """
    grouped = defaultdict(list)
    
    for row in data:
        # 只处理未入账的数据
        if row['入账'] != '是':
            key = (row['出库日期'], row['出库对象'])
            grouped[key].append(row)
    
    return grouped

def group_by_product(items):
    """
    按产品规格分组,计算汇总
    返回: {规格: {'件数': x, '净重列表': [], '总净重': x}}
    """
    products = defaultdict(lambda: {'件数': 0, '净重列表': [], '总净重': 0.0})
    
    for item in items:
        spec = item['规格']
        net_weight = float(item['净重']) if item['净重'] else 0.0
        
        products[spec]['件数'] += 1
        products[spec]['净重列表'].append(net_weight)
        products[spec]['总净重'] += net_weight
    
    return products

def create_simple_invoice(wb, date_str, customer, items, invoice_no):
    """
    创建简单版销售清单(基于TemplateSheet)
    """
    # 复制模板
    template_ws = wb['TemplateSheet']
    sheet_name = f"销货清单_{customer}_{date_str}_{invoice_no}_简单版"
    
    # 创建新工作表
    new_ws = wb.copy_worksheet(template_ws)
    new_ws.title = sheet_name[:31]  # Excel工作表名称限制31字符
    
    # 填充数据
    # 客户名称 (C3)
    new_ws['B3'] = f"客户: {customer}"
    
    # 开单日期 (F3)
    new_ws['F3'] = f" 开单日期: {date_str}"
    
    # 单号 (I2)
    new_ws['I2'] = f"NO {invoice_no}"
    
    # 按产品分组
    products = group_by_product(items)
    
    # 填充产品明细 (从第5行开始)
    row_idx = 5
    total_amount = 0.0
    
    for spec, info in products.items():
        new_ws.cell(row_idx, 1).value = spec  # 产品名称
        new_ws.cell(row_idx, 2).value = info['件数']  # 件数
        new_ws.cell(row_idx, 3).value = round(info['总净重'], 2)  # 总重量
        # 单价和金额需要手动填写
        new_ws.cell(row_idx, 4).value = ""  # 单价
        new_ws.cell(row_idx, 5).value = ""  # 金额
        
        # 明细净重
        detail_str = ", ".join([str(round(w, 2)) for w in info['净重列表']])
        new_ws.cell(row_idx, 6).value = f"明细净重(kg): {detail_str}"
        
        row_idx += 1
    
    print(f"  ✓ 创建简单版销售清单: {new_ws.title}")
    return new_ws

def create_detailed_invoice(wb, date_str, customer, items, invoice_no):
    """
    创建详细版销售清单(基于pasted_content.txt的格式)
    """
    sheet_name = f"销货清单_{customer}_{date_str}_{invoice_no}_详细版"
    
    # 创建新工作表
    new_ws = wb.create_sheet(title=sheet_name[:31])
    
    # 设置列宽
    new_ws.column_dimensions['A'].width = 20
    new_ws.column_dimensions['B'].width = 12
    new_ws.column_dimensions['C'].width = 15
    new_ws.column_dimensions['D'].width = 12
    new_ws.column_dimensions['E'].width = 15
    
    # 定义样式
    title_font = Font(name='宋体', size=16, bold=True)
    header_font = Font(name='宋体', size=12, bold=True)
    normal_font = Font(name='宋体', size=11)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    row_idx = 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = "东阳市欧亚金银丝有限公司"
    cell.font = title_font
    cell.alignment = center_align
    
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = "销货清单"
    cell.font = header_font
    cell.alignment = center_align
    
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:C{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = f"客户: {customer}"
    cell.font = normal_font
    cell.alignment = left_align
    
    new_ws.merge_cells(f'D{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 4)
    cell.value = f"No. {invoice_no}"
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = f"开单日期: {date_str}"
    cell.font = normal_font
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # 表头
    row_idx += 1
    headers = ['产品名称', '件数', '总重量(kg)', '单价(元)', '金额(元)']
    for col_idx, header in enumerate(headers, start=1):
        cell = new_ws.cell(row_idx, col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 按产品分组
    products = group_by_product(items)
    
    # 填充产品明细
    total_pieces = 0
    total_weight = 0.0
    
    for spec, info in products.items():
        row_idx += 1
        
        # 产品行
        new_ws.cell(row_idx, 1).value = spec
        new_ws.cell(row_idx, 1).font = normal_font
        new_ws.cell(row_idx, 1).alignment = center_align
        new_ws.cell(row_idx, 1).border = thin_border
        
        new_ws.cell(row_idx, 2).value = info['件数']
        new_ws.cell(row_idx, 2).font = normal_font
        new_ws.cell(row_idx, 2).alignment = center_align
        new_ws.cell(row_idx, 2).border = thin_border
        
        new_ws.cell(row_idx, 3).value = round(info['总净重'], 2)
        new_ws.cell(row_idx, 3).font = normal_font
        new_ws.cell(row_idx, 3).alignment = center_align
        new_ws.cell(row_idx, 3).border = thin_border
        
        # 单价和金额留空,需要手动填写
        new_ws.cell(row_idx, 4).value = ""
        new_ws.cell(row_idx, 4).font = normal_font
        new_ws.cell(row_idx, 4).alignment = center_align
        new_ws.cell(row_idx, 4).border = thin_border
        
        new_ws.cell(row_idx, 5).value = ""
        new_ws.cell(row_idx, 5).font = normal_font
        new_ws.cell(row_idx, 5).alignment = center_align
        new_ws.cell(row_idx, 5).border = thin_border
        
        # 明细净重
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        detail_str = ", ".join([str(round(w, 2)) for w in info['净重列表']])
        cell = new_ws.cell(row_idx, 1)
        cell.value = f"明细净重(kg): {detail_str}"
        cell.font = Font(name='宋体', size=10)
        cell.alignment = left_align
        cell.border = thin_border
        
        total_pieces += info['件数']
        total_weight += info['总净重']
    
    # 汇总
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = f"汇总: 总件数 {total_pieces}箱    总重量 {round(total_weight, 2)}kg"
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # 金额汇总(留空)
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = "合计金额(大写): "
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = "合计金额(小写): ¥"
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    
    # 备注
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = "备注: 1. 建议用户试样,如有质量问题,请在3日内退回。2. 如果发生法律纠纷,由东阳市人民法院管辖。"
    cell.font = Font(name='宋体', size=9)
    cell.alignment = left_align
    
    row_idx += 1
    new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
    cell = new_ws.cell(row_idx, 1)
    cell.value = "手机: 18606833896, 18606886823  电话: 0579-86985290  传真: 0579-86985471"
    cell.font = Font(name='宋体', size=9)
    cell.alignment = center_align
    
    print(f"  ✓ 创建详细版销售清单: {new_ws.title}")
    return new_ws

def mark_as_recorded(ws, row_indices):
    """
    在入账列标记"是"
    """
    for row_idx in row_indices:
        ws.cell(row_idx, 9).value = "是"

def generate_invoices(wb):
    """
    生成销售清单
    """
    print("\n正在生成销售清单...")
    
    # 读取数据
    ws = wb['BondDataSheet']
    data = read_bond_data(ws)
    
    # 按日期和客户分组
    grouped = group_data_by_date_and_customer(data)
    
    if not grouped:
        print("  没有需要生成销售清单的数据(所有数据都已入账)")
        return
    
    invoice_counter = 1
    
    for (date_obj, customer), items in grouped.items():
        date_str = date_obj.strftime('%Y-%m-%d')
        invoice_no = f"{invoice_counter:05d}"
        
        print(f"\n处理: {date_str} - {customer} ({len(items)}条记录)")
        
        # 生成简单版
        create_simple_invoice(wb, date_str, customer, items, invoice_no)
        
        # 生成详细版
        create_detailed_invoice(wb, date_str, customer, items, invoice_no)
        
        # 标记为已入账
        row_indices = [item['row_idx'] for item in items]
        mark_as_recorded(ws, row_indices)
        
        invoice_counter += 1
    
    print(f"\n✓ 共生成 {len(grouped)} 组销售清单")

def main():
    """
    主函数
    """
    input_file = '库存tmep.xlsx'
    output_file = '库存_改进版.xlsx'
    
    print("=" * 60)
    print("库存表改进脚本")
    print("=" * 60)
    
    # 加载工作簿
    print(f"\n正在加载文件: {input_file}")
    wb = openpyxl.load_workbook(input_file)
    
    # 1. 改进BondDataTable
    improve_bond_data_table(wb)
    
    # 2. 生成销售清单
    generate_invoices(wb)
    
    # 保存文件
    print(f"\n正在保存文件: {output_file}")
    wb.save(output_file)
    
    print("\n" + "=" * 60)
    print("✓ 所有操作完成!")
    print("=" * 60)
    print(f"\n输出文件: {output_file}")
    print("\n说明:")
    print("1. BondDataTable已优化,新增行会自动填充序号和日期")
    print("2. 已为所有未入账的数据生成销售清单(简单版+详细版)")
    print("3. 已生成清单的数据在'入账'列标记为'是'")
    print("4. 销售清单中的单价和金额需要手动填写")

if __name__ == '__main__':
    main()
