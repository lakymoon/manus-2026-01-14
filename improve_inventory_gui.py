#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
库存表改进脚本 - GUI版本
功能:
1. 修改BondDataTable,实现序号和日期自动填充
2. 根据BondDataTable自动生成销售清单(两种模板)
3. 提供图形界面,支持文件选择和进度显示
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import sys
import traceback

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("库存表自动化管理系统 v1.0")
        self.root.geometry("700x500")
        self.root.resizable(False, False)
        
        # 设置图标(如果有的话)
        try:
            self.root.iconbitmap('icon.ico')
        except:
            pass
        
        self.input_file = None
        self.output_file = None
        
        self.create_widgets()
    
    def create_widgets(self):
        # 标题
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=60)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(
            title_frame, 
            text="📊 库存表自动化管理系统",
            font=("微软雅黑", 18, "bold"),
            fg="white",
            bg="#2c3e50"
        )
        title_label.pack(pady=15)
        
        # 主内容区
        main_frame = tk.Frame(self.root, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文件选择区域
        file_frame = tk.LabelFrame(main_frame, text="📁 选择文件", font=("微软雅黑", 11, "bold"), padx=10, pady=10)
        file_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 输入文件
        input_frame = tk.Frame(file_frame)
        input_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(input_frame, text="输入文件:", font=("微软雅黑", 10), width=10, anchor='w').pack(side=tk.LEFT)
        self.input_entry = tk.Entry(input_frame, font=("微软雅黑", 9), state='readonly')
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Button(
            input_frame, 
            text="浏览...", 
            command=self.select_input_file,
            font=("微软雅黑", 9),
            bg="#3498db",
            fg="white",
            relief=tk.FLAT,
            padx=15,
            cursor="hand2"
        ).pack(side=tk.LEFT)
        
        # 输出文件
        output_frame = tk.Frame(file_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(output_frame, text="输出文件:", font=("微软雅黑", 10), width=10, anchor='w').pack(side=tk.LEFT)
        self.output_entry = tk.Entry(output_frame, font=("微软雅黑", 9), state='readonly')
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Button(
            output_frame, 
            text="浏览...", 
            command=self.select_output_file,
            font=("微软雅黑", 9),
            bg="#3498db",
            fg="white",
            relief=tk.FLAT,
            padx=15,
            cursor="hand2"
        ).pack(side=tk.LEFT)
        
        # 日志区域
        log_frame = tk.LabelFrame(main_frame, text="📝 运行日志", font=("微软雅黑", 11, "bold"), padx=10, pady=10)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, 
            font=("Consolas", 9),
            height=12,
            wrap=tk.WORD,
            bg="#f8f9fa",
            relief=tk.FLAT
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 按钮区域
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        self.run_button = tk.Button(
            button_frame,
            text="🚀 开始处理",
            command=self.run_process,
            font=("微软雅黑", 12, "bold"),
            bg="#27ae60",
            fg="white",
            relief=tk.FLAT,
            padx=30,
            pady=10,
            cursor="hand2"
        )
        self.run_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 5))
        
        tk.Button(
            button_frame,
            text="❌ 退出",
            command=self.root.quit,
            font=("微软雅黑", 12, "bold"),
            bg="#e74c3c",
            fg="white",
            relief=tk.FLAT,
            padx=30,
            pady=10,
            cursor="hand2"
        ).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(5, 0))
        
        # 初始日志
        self.log("欢迎使用库存表自动化管理系统!")
        self.log("请选择输入文件(库存tmep.xlsx)开始处理...")
        self.log("-" * 60)
    
    def log(self, message):
        """添加日志"""
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def select_input_file(self):
        """选择输入文件"""
        filename = filedialog.askopenfilename(
            title="选择库存Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.input_file = filename
            self.input_entry.config(state='normal')
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, filename)
            self.input_entry.config(state='readonly')
            
            # 自动设置输出文件名
            dir_name = os.path.dirname(filename)
            base_name = os.path.basename(filename)
            name, ext = os.path.splitext(base_name)
            output_name = f"{name}_改进版{ext}"
            self.output_file = os.path.join(dir_name, output_name)
            
            self.output_entry.config(state='normal')
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, self.output_file)
            self.output_entry.config(state='readonly')
            
            self.log(f"✓ 已选择输入文件: {os.path.basename(filename)}")
    
    def select_output_file(self):
        """选择输出文件"""
        filename = filedialog.asksaveasfilename(
            title="保存输出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.output_file = filename
            self.output_entry.config(state='normal')
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, filename)
            self.output_entry.config(state='readonly')
            
            self.log(f"✓ 已设置输出文件: {os.path.basename(filename)}")
    
    def run_process(self):
        """运行处理流程"""
        if not self.input_file:
            messagebox.showwarning("警告", "请先选择输入文件!")
            return
        
        if not self.output_file:
            messagebox.showwarning("警告", "请先设置输出文件!")
            return
        
        try:
            self.run_button.config(state='disabled', text="处理中...")
            self.log("\n" + "=" * 60)
            self.log("开始处理...")
            self.log("=" * 60)
            
            # 加载工作簿
            self.log(f"\n正在加载文件: {os.path.basename(self.input_file)}")
            wb = openpyxl.load_workbook(self.input_file)
            
            # 改进BondDataTable
            self.improve_bond_data_table(wb)
            
            # 生成销售清单
            self.generate_invoices(wb)
            
            # 保存文件
            self.log(f"\n正在保存文件: {os.path.basename(self.output_file)}")
            wb.save(self.output_file)
            
            self.log("\n" + "=" * 60)
            self.log("✓ 所有操作完成!")
            self.log("=" * 60)
            self.log(f"\n输出文件: {self.output_file}")
            
            messagebox.showinfo("成功", f"处理完成!\n\n输出文件:\n{self.output_file}")
            
        except Exception as e:
            error_msg = f"错误: {str(e)}\n\n{traceback.format_exc()}"
            self.log(f"\n❌ 处理失败:\n{error_msg}")
            messagebox.showerror("错误", f"处理失败:\n{str(e)}")
        
        finally:
            self.run_button.config(state='normal', text="🚀 开始处理")
    
    def improve_bond_data_table(self, wb):
        """改进BondDataTable"""
        ws = wb['BondDataSheet']
        self.log("\n正在改进BondDataTable...")
        
        for row_idx in range(2, ws.max_row + 1):
            # 序号列
            seq_cell = ws.cell(row_idx, 1)
            if seq_cell.value is None or (isinstance(seq_cell.value, str) and seq_cell.value.startswith('=')):
                seq_cell.value = f'=ROW(BondDataTable[[#This Row],[序号]])-1'
            
            # 出库日期列
            date_cell = ws.cell(row_idx, 2)
            if date_cell.value is None:
                date_cell.value = date.today()
                date_cell.number_format = 'YYYY-MM-DD'
            elif isinstance(date_cell.value, datetime):
                date_cell.number_format = 'YYYY-MM-DD'
            
            # 净重列
            net_weight_cell = ws.cell(row_idx, 7)
            if net_weight_cell.value is None or (isinstance(net_weight_cell.value, str) and net_weight_cell.value.startswith('=')):
                net_weight_cell.value = f'=BondDataTable[[#This Row],[毛重]]-BondDataTable[[#This Row],[除皮]]'
        
        self.log("  ✓ BondDataTable改进完成")
    
    def read_bond_data(self, ws):
        """从BondDataSheet读取数据"""
        data = []
        
        for row_idx in range(2, ws.max_row + 1):
            net_weight_cell = ws.cell(row_idx, 7)
            if net_weight_cell.data_type == 'f':
                try:
                    net_weight = float(ws.cell(row_idx, 5).value) - float(ws.cell(row_idx, 6).value)
                except:
                    net_weight = 0.0
            else:
                net_weight = net_weight_cell.value
            
            row_data = {
                '序号': ws.cell(row_idx, 1).value,
                '出库日期': ws.cell(row_idx, 2).value,
                '规格': ws.cell(row_idx, 3).value,
                '个数': ws.cell(row_idx, 4).value,
                '毛重': ws.cell(row_idx, 5).value,
                '除皮': ws.cell(row_idx, 6).value,
                '净重': net_weight,
                '出库对象': ws.cell(row_idx, 8).value,
                '入账': ws.cell(row_idx, 9).value,
                '备注': ws.cell(row_idx, 10).value,
                'row_idx': row_idx
            }
            
            if row_data['出库日期'] is None or row_data['出库对象'] is None:
                continue
            
            if isinstance(row_data['出库日期'], datetime):
                row_data['出库日期'] = row_data['出库日期'].date()
            
            data.append(row_data)
        
        return data
    
    def group_data_by_date_and_customer(self, data):
        """按出库日期和出库对象分组"""
        grouped = defaultdict(list)
        
        for row in data:
            if row['入账'] != '是':
                key = (row['出库日期'], row['出库对象'])
                grouped[key].append(row)
        
        return grouped
    
    def group_by_product(self, items):
        """按产品规格分组"""
        products = defaultdict(lambda: {'件数': 0, '净重列表': [], '总净重': 0.0})
        
        for item in items:
            spec = item['规格']
            net_weight = float(item['净重']) if item['净重'] else 0.0
            
            products[spec]['件数'] += 1
            products[spec]['净重列表'].append(net_weight)
            products[spec]['总净重'] += net_weight
        
        return products
    
    def create_simple_invoice(self, wb, date_str, customer, items, invoice_no):
        """创建简单版销售清单"""
        template_ws = wb['TemplateSheet']
        sheet_name = f"销货清单_{customer}_{date_str}_{invoice_no}_简单版"
        
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = sheet_name[:31]
        
        new_ws['B3'] = f"客户: {customer}"
        new_ws['F3'] = f" 开单日期: {date_str}"
        new_ws['I2'] = f"NO {invoice_no}"
        
        products = self.group_by_product(items)
        
        row_idx = 5
        for spec, info in products.items():
            new_ws.cell(row_idx, 1).value = spec
            new_ws.cell(row_idx, 2).value = info['件数']
            new_ws.cell(row_idx, 3).value = round(info['总净重'], 2)
            new_ws.cell(row_idx, 4).value = ""
            new_ws.cell(row_idx, 5).value = ""
            
            detail_str = ", ".join([str(round(w, 2)) for w in info['净重列表']])
            new_ws.cell(row_idx, 6).value = f"明细净重(kg): {detail_str}"
            
            row_idx += 1
        
        self.log(f"  ✓ 创建简单版: {customer}")
        return new_ws
    
    def create_detailed_invoice(self, wb, date_str, customer, items, invoice_no):
        """创建详细版销售清单"""
        sheet_name = f"销货清单_{customer}_{date_str}_{invoice_no}_详细版"
        new_ws = wb.create_sheet(title=sheet_name[:31])
        
        new_ws.column_dimensions['A'].width = 20
        new_ws.column_dimensions['B'].width = 12
        new_ws.column_dimensions['C'].width = 15
        new_ws.column_dimensions['D'].width = 12
        new_ws.column_dimensions['E'].width = 15
        
        title_font = Font(name='宋体', size=16, bold=True)
        header_font = Font(name='宋体', size=12, bold=True)
        normal_font = Font(name='宋体', size=11)
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_idx = 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = "东阳市欧亚金银丝有限公司"
        cell.font = title_font
        cell.alignment = center_align
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = "销货清单"
        cell.font = header_font
        cell.alignment = center_align
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:C{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = f"客户: {customer}"
        cell.font = normal_font
        cell.alignment = left_align
        
        new_ws.merge_cells(f'D{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 4)
        cell.value = f"No. {invoice_no}"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = f"开单日期: {date_str}"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        row_idx += 1
        headers = ['产品名称', '件数', '总重量(kg)', '单价(元)', '金额(元)']
        for col_idx, header in enumerate(headers, start=1):
            cell = new_ws.cell(row_idx, col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        products = self.group_by_product(items)
        
        total_pieces = 0
        total_weight = 0.0
        
        for spec, info in products.items():
            row_idx += 1
            
            new_ws.cell(row_idx, 1).value = spec
            new_ws.cell(row_idx, 1).font = normal_font
            new_ws.cell(row_idx, 1).alignment = center_align
            new_ws.cell(row_idx, 1).border = thin_border
            
            new_ws.cell(row_idx, 2).value = info['件数']
            new_ws.cell(row_idx, 2).font = normal_font
            new_ws.cell(row_idx, 2).alignment = center_align
            new_ws.cell(row_idx, 2).border = thin_border
            
            new_ws.cell(row_idx, 3).value = round(info['总净重'], 2)
            new_ws.cell(row_idx, 3).font = normal_font
            new_ws.cell(row_idx, 3).alignment = center_align
            new_ws.cell(row_idx, 3).border = thin_border
            
            new_ws.cell(row_idx, 4).value = ""
            new_ws.cell(row_idx, 4).font = normal_font
            new_ws.cell(row_idx, 4).alignment = center_align
            new_ws.cell(row_idx, 4).border = thin_border
            
            new_ws.cell(row_idx, 5).value = ""
            new_ws.cell(row_idx, 5).font = normal_font
            new_ws.cell(row_idx, 5).alignment = center_align
            new_ws.cell(row_idx, 5).border = thin_border
            
            row_idx += 1
            new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
            detail_str = ", ".join([str(round(w, 2)) for w in info['净重列表']])
            cell = new_ws.cell(row_idx, 1)
            cell.value = f"明细净重(kg): {detail_str}"
            cell.font = Font(name='宋体', size=10)
            cell.alignment = left_align
            cell.border = thin_border
            
            total_pieces += info['件数']
            total_weight += info['总净重']
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = f"汇总: 总件数 {total_pieces}箱    总重量 {round(total_weight, 2)}kg"
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = "合计金额(大写): "
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = "合计金额(小写): ¥"
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = "备注: 1. 建议用户试样,如有质量问题,请在3日内退回。2. 如果发生法律纠纷,由东阳市人民法院管辖。"
        cell.font = Font(name='宋体', size=9)
        cell.alignment = left_align
        
        row_idx += 1
        new_ws.merge_cells(f'A{row_idx}:E{row_idx}')
        cell = new_ws.cell(row_idx, 1)
        cell.value = "手机: 18606833896, 18606886823  电话: 0579-86985290  传真: 0579-86985471"
        cell.font = Font(name='宋体', size=9)
        cell.alignment = center_align
        
        self.log(f"  ✓ 创建详细版: {customer}")
        return new_ws
    
    def mark_as_recorded(self, ws, row_indices):
        """在入账列标记'是'"""
        for row_idx in row_indices:
            ws.cell(row_idx, 9).value = "是"
    
    def generate_invoices(self, wb):
        """生成销售清单"""
        self.log("\n正在生成销售清单...")
        
        ws = wb['BondDataSheet']
        data = self.read_bond_data(ws)
        
        grouped = self.group_data_by_date_and_customer(data)
        
        if not grouped:
            self.log("  没有需要生成销售清单的数据(所有数据都已入账)")
            return
        
        invoice_counter = 1
        
        for (date_obj, customer), items in grouped.items():
            date_str = date_obj.strftime('%Y-%m-%d')
            invoice_no = f"{invoice_counter:05d}"
            
            self.log(f"\n处理: {date_str} - {customer} ({len(items)}条记录)")
            
            self.create_simple_invoice(wb, date_str, customer, items, invoice_no)
            self.create_detailed_invoice(wb, date_str, customer, items, invoice_no)
            
            row_indices = [item['row_idx'] for item in items]
            self.mark_as_recorded(ws, row_indices)
            
            invoice_counter += 1
        
        self.log(f"\n✓ 共生成 {len(grouped)} 组销售清单(简单版+详细版)")

def main():
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()
